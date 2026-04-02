# Reporting Module for AWS Automated Access Review
import csv
import boto3
import os
from datetime import datetime
from io import StringIO, BytesIO

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_report(findings, format_type="csv", local_mode=False):
    """
    Merges findings into a report (CSV or XLSX), uploads to S3 or saves locally.

    Args:
        findings: List of finding dictionaries
        format_type: 'csv' or 'xlsx' (default: 'csv')
        local_mode: If True, save to local 'reports' folder instead of S3

    Returns:
        Tuple of (file_path, url_or_file_path)
    """
    # Check if we should use local mode
    bucket_name = os.environ.get("REPORT_BUCKET")
    if local_mode or not bucket_name:
        return _generate_local_report(findings, format_type)

    timestamp = datetime.now().isoformat()
    date_str = datetime.now().strftime("%Y-%m-%d")

    if format_type.lower() == "xlsx" and HAS_OPENPYXL:
        return _generate_xlsx_report(findings, bucket_name, date_str, timestamp)
    else:
        return _generate_csv_report(findings, bucket_name, date_str, timestamp)


def _generate_local_report(findings, format_type):
    """Generate report locally in 'reports' folder."""
    # Create reports directory if it doesn't exist
    reports_dir = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "reports"
    )
    os.makedirs(reports_dir, exist_ok=True)

    timestamp = datetime.now().isoformat()
    date_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    if format_type.lower() == "xlsx" and HAS_OPENPYXL:
        return _save_xlsx_locally(findings, reports_dir, date_str, timestamp)
    else:
        return _save_csv_locally(findings, reports_dir, date_str, timestamp)


def _save_csv_locally(findings, reports_dir, date_str, timestamp):
    """Save CSV report locally."""
    filename = f"access-review-report-{date_str}.csv"
    filepath = os.path.join(reports_dir, filename)

    headers = [
        "Timestamp",
        "ResourceID",
        "ResourceType",
        "Service",
        "Severity",
        "Finding",
        "Recommendation",
    ]

    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers)
        writer.writeheader()

        for finding in findings:
            writer.writerow(
                {
                    "Timestamp": timestamp,
                    "ResourceID": finding.get("resource_id", "N/A"),
                    "ResourceType": finding.get("resource_type", "N/A"),
                    "Service": finding.get("service", "N/A"),
                    "Severity": finding.get("severity", "N/A"),
                    "Finding": finding.get("finding", "N/A"),
                    "Recommendation": finding.get("recommendation", "N/A"),
                }
            )

    print(f"CSV report saved locally: {filepath}")
    return filepath, filepath


def _save_xlsx_locally(findings, reports_dir, date_str, timestamp):
    """Save XLSX report locally."""
    filename = f"access-review-report-{date_str}.xlsx"
    filepath = os.path.join(reports_dir, filename)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Access Review Findings"

    # Write headers
    headers = [
        "Timestamp",
        "Resource ID",
        "Resource Type",
        "Service",
        "Severity",
        "Finding",
        "Recommendation",
    ]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    # Write findings
    for row_num, f in enumerate(findings, 2):
        worksheet.cell(row=row_num, column=1, value=timestamp)
        worksheet.cell(row=row_num, column=2, value=f.get("resource_id", "N/A"))
        worksheet.cell(row=row_num, column=3, value=f.get("resource_type", "N/A"))
        worksheet.cell(row=row_num, column=4, value=f.get("service", "N/A"))

        # Color code severity
        severity = f.get("severity", "N/A")
        severity_cell = worksheet.cell(row=row_num, column=5, value=severity)
        if severity == "CRITICAL":
            severity_cell.fill = openpyxl.styles.PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )
        elif severity == "HIGH":
            severity_cell.fill = openpyxl.styles.PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )
        elif severity == "MEDIUM":
            severity_cell.fill = openpyxl.styles.PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

        worksheet.cell(row=row_num, column=6, value=f.get("finding", "N/A"))
        worksheet.cell(row=row_num, column=7, value=f.get("recommendation", "N/A"))

    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(filepath)
    print(f"XLSX report saved locally: {filepath}")
    return filepath, filepath


def _generate_csv_report(findings, bucket_name, date_str, timestamp):
    """Generate CSV report."""
    # Create CSV in memory
    csv_buffer = StringIO()
    headers = [
        "Timestamp",
        "ResourceID",
        "ResourceType",
        "Service",
        "Severity",
        "Finding",
        "Recommendation",
    ]
    writer = csv.DictWriter(csv_buffer, fieldnames=headers)
    writer.writeheader()

    for f in findings:
        writer.writerow(
            {
                "Timestamp": timestamp,
                "ResourceID": f.get("resource_id", "N/A"),
                "ResourceType": f.get("resource_type", "N/A"),
                "Service": f.get("service", "N/A"),
                "Severity": f.get("severity", "N/A"),
                "Finding": f.get("finding", "N/A"),
                "Recommendation": f.get("recommendation", "N/A"),
            }
        )

    # Upload to S3
    s3 = boto3.client("s3")
    s3_key = f"reports/{date_str}/access-review-report.csv"

    try:
        s3.put_object(
            Bucket=bucket_name,
            Key=s3_key,
            Body=csv_buffer.getvalue(),
            ContentType="text/csv",
        )

        # Generate pre-signed URL (7 days)
        url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": bucket_name, "Key": s3_key},
            ExpiresIn=604800,
        )

        return s3_key, url
    except Exception as e:
        print(f"Error uploading report to S3: {e}")
        return None, None


def _generate_xlsx_report(findings, bucket_name, date_str, timestamp):
    """Generate XLSX report."""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Access Review Findings"

    # Write headers
    headers = [
        "Timestamp",
        "Resource ID",
        "Resource Type",
        "Service",
        "Severity",
        "Finding",
        "Recommendation",
    ]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    # Write findings
    for row_num, f in enumerate(findings, 2):
        worksheet.cell(row=row_num, column=1, value=timestamp)
        worksheet.cell(row=row_num, column=2, value=f.get("resource_id", "N/A"))
        worksheet.cell(row=row_num, column=3, value=f.get("resource_type", "N/A"))
        worksheet.cell(row=row_num, column=4, value=f.get("service", "N/A"))

        # Color code severity
        severity = f.get("severity", "N/A")
        severity_cell = worksheet.cell(row=row_num, column=5, value=severity)
        if severity == "CRITICAL":
            severity_cell.fill = openpyxl.styles.PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid"
            )
        elif severity == "HIGH":
            severity_cell.fill = openpyxl.styles.PatternFill(
                start_color="FFA500", end_color="FFA500", fill_type="solid"
            )
        elif severity == "MEDIUM":
            severity_cell.fill = openpyxl.styles.PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

        worksheet.cell(row=row_num, column=6, value=f.get("finding", "N/A"))
        worksheet.cell(row=row_num, column=7, value=f.get("recommendation", "N/A"))

    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column].width = adjusted_width

    # Save to buffer
    xlsx_buffer = BytesIO()
    workbook.save(xlsx_buffer)
    xlsx_buffer.seek(0)

    # Upload to S3
    s3 = boto3.client("s3")
    s3_key = f"reports/{date_str}/access-review-report.xlsx"

    try:
        s3.put_object(
            Bucket=bucket_name,
            Key=s3_key,
            Body=xlsx_buffer.getvalue(),
            ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Generate pre-signed URL (7 days)
        url = s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": bucket_name, "Key": s3_key},
            ExpiresIn=604800,
        )

        return s3_key, url
    except Exception as e:
        print(f"Error uploading report to S3: {e}")
        return None, None
